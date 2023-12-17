import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger

class ManipulaPlanilhas:
    
    def __init__(self, arquivo_xlsx: str)-> None:
        """
        Inicializa a instância da classe ManipulaPlanilha.

        Args:
            arquivo_xlsx (str): O caminho do arquivo XLSX a ser manipulado.
        """
        self.arquivo_xlsx = arquivo_xlsx

        if os.path.exists(arquivo_xlsx):
            self.wb = load_workbook(filename=arquivo_xlsx)
        else:
            self.wb = Workbook()

    def preenche_planilha(self, planilha: str , dados: dict) -> None:
        """
        Adiciona dados a uma planilha e aplica estilos às células.

        Args:
            planilha (str): O nome da planilha onde os dados serão adicionados.
            dados (dict): Um dicionário contendo os dados a serem adicionados à planilha.
                Exemplo:
                {
                    'Nome': ['João', 'Maria', 'Carlos', 'Ana'],
                    'Idade': [25, 30, 22, 28],
                    'Cidade': ['São Paulo', 'Rio de Janeiro', 'Belo Horizonte', 'Porto Alegre'],
                    'Salário': [50000, 60000, 45000, 55000]
                }
        """

        if planilha in self.wb.sheetnames:
            ws = self.wb[planilha]
        else:
            ws = self.wb.create_sheet(title=planilha, index=0)
        
        i_coluna = 1

        for coluna, valores in dados.items():
            # Define o cabeçalho (linha 1) da coluna com o seu respectivo valor
            celula_coluna = ws.cell(row=1, column=i_coluna, value=coluna)

            # Aplica estilos ao cabeçalho
            celula_coluna.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            celula_coluna.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            celula_coluna.alignment = Alignment(horizontal='center')
            celula_coluna.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Indice da linha inicial em que os dados serão preenchidos na planilha
            i_linha = 2

            # Itera sobre os dados referentes à coluna atual
            for dado in valores:
                # Preenche o campo referente à coluna e linha atual com o seu respectivo dado
                celula_dado = ws.cell(row=i_linha, column=i_coluna, value=dado)

                # Aplica estilos diretamente à célula de dado
                celula_dado.font = Font(name='Arial', size=11, bold=False)
                celula_dado.alignment = Alignment(horizontal='left')
                celula_dado.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


                # Incrementa 1 ao índice da linha para a próxima iteração
                i_linha += 1

            # Incrementa 1 ao índice da coluna para a próxima iteração
            i_coluna += 1

        # Cabeçalho fixo
        ws.freeze_panes = 'A2'

        # ---- SALVA A PLANILHA ----
        self.wb.save(filename=self.arquivo_xlsx)

    def obter_index_coluna(self, planilha: str, nome_coluna: str) -> int:
        """
        Obtém o índice da coluna pelo seu nome.

        Args:
            planilha (str): O nome da planilha onde a coluna está localizada.
            nome_coluna (str): O nome da coluna.

        Returns:
            int: O índice da coluna.
        """

        # Define a planilha
        ws = self.wb[planilha]

        # Itera sobre as colunas presentes na planilha
        for coluna in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):

            # Itera sobe a celula da coluna atual
            for celula in coluna:

                # Verifica se o valor da celula é igual ao nome da coluna busca
                if celula.value == nome_coluna:

                    # Retorna o indice da coluna
                    return cell.column
        return None

    def preenche_coluna(self, planilha:str, coluna_destino, coluna_referencia, dados: dict) -> None:
        """
        Preenche uma coluna com informações usando como base valores presentes em uma coluna de 
        referencia

        Args:
            planilha (str): O nome da planilha onde os dados serão adicionados.
            coluna_destino (str): O nome da coluna em que os dados deverão ser inseridos.
            coluna_refenrecia (str): O nome da coluna que deverá ser utilizada como referência.
            dados (dict): Um dicionário contendo os dados a serem adicionados à planilha.
                Exemplo:
                {
                    'valor presente na coluna referencia': 'valor a ser inserido',
                    'valor presente na coluna referencia': 'valor a ser inserido',
                }
        """

        # Define a planilha
        ws = self.wb[planilha]

        # Obtém índice das colunas de referência
        coluna_referencia_index = self.obter_index_coluna(
            planilha = planilha, 
            nome_coluna = coluna_referencia
        )

        # Obtém índice das colunas de referência
        coluna_destino_index = self.obter_index_coluna(
            planilha = planilha,
            nome_coluna = coluna_destino
        )

        # Itera sobre as linhas de dados da planilha
        for linha in range(2, ws.max_row + 1):

            # Define o valor da linha atual na coluna referência
            valor_referencia = ws.cell(row=linha, column=coluna_referencia_index).value

            # Verifica se o valor referência está presente em uma das chaves do dict de dados
            if valor_referencia in dados.keys():

                # Define o valor que será inserido
                valor_inserir = dados[valor_referencia]

                # Insere o valor na planilha
                celula = ws.cell(
                    row=linha,
                    column=coluna_destino_index,
                    value=valor_inserir
                )

                # Aplica estilos diretamente à célula
                celula.font = Font(
                    name='Arial',
                    size=11,
                    bold=False
                )

                celula.alignment = Alignment(
                    horizontal='left'
                )
                
                celula.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

        # Salva as alterações realizadas
        self.wb.save(filename=self.arquivo_xlsx)



if __name__ == '__main__':

    # ---- CAMINHOS ----
    DIRETORIO_PLANILHA = r'C:\Projetos Pessoais\web-scraping-mercado\assets\planilhas'
    PLANILHA_1 = os.path.join(DIRETORIO_PLANILHA, '1.xlsx')

    # ---- DADOS DE ENTRADA ----
    dados = {
        'Nome': ['João', 'Maria', 'Carlos', 'Ana'],
        'Idade': [25, 30, 22, 28],
        'Cidade': ['São Paulo', 'Rio de Janeiro', 'Belo Horizonte', 'Porto Alegre'],
        'Salário': [50000, 60000, 45000, 55000],
        'Retorno' : []
    }

    manipula_planilhas = ManipulaPlanilhas(arquivo_xlsx=PLANILHA_1)

    manipula_planilhas.preenche_planilha(
        planilha='Principal',
        dados=dados
    )

    i = manipula_planilhas.obter_index_coluna(
        ws = 'Principal',
        nome_coluna='Nome'
    )

    logger.info(i)
 
    manipula_planilhas.preenche_coluna(
        planilha='Principal',
        coluna_referencia='Nome',
        coluna_destino='Retorno',
        dados={
            'João' : 'Sucesso',
            'Maria' : 'Erro',
            'Carlos': 'Erro',
            'Ana' : 'Sucesso' 
        }
    )
